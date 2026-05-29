"""
Main routes for serving HTML pages.
"""
import io
import os
import zipfile
from datetime import datetime
from urllib.parse import quote

from flask import Blueprint, current_app, flash, g, send_file, send_from_directory, render_template, redirect, request, url_for
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from sqlalchemy.orm import joinedload

from app.models import Initiative, FacilityAllocation, Rebate, User, UserRole
from app.utils.decorators import get_current_user, login_required
from app.utils.email import (
    is_graph_mail_configured,
    send_initiative_approved_notification,
    send_initiative_creator_notification,
    send_initiative_review_notification,
    send_weekly_review_reminder,
)
from app.utils.runtime_settings import apply_static_settings, load_static_settings, save_static_settings
from app.utils.timezone import now_eastern

main_bp = Blueprint('main', __name__)

_REVIEW_NOTIFICATION_DEFAULT_EMAIL = 'procurementdatateam@montefiore.org'

_FINANCE_ALLOWED_ENDPOINTS = {
    'main.rebate_extraction',
    'main.rebate_extraction_export',
    'main.logout',
    'main.send_static',
    'static',
}


@main_bp.app_context_processor
def inject_template_user():
    """Expose the signed-in user object and current environment to Jinja templates."""
    user = getattr(g, 'current_user', None)
    if user is None:
        try:
            user = get_current_user()
        except Exception:
            user = None
    env_name = os.getenv('ENVIRONMENT', os.getenv('FLASK_ENV', 'production')).lower()
    return {
        'template_current_user': user,
        'template_is_readonly_user': _is_read_only_user(user),
        'template_role_request_mailto': _build_role_request_mailto(user),
        'template_environment': env_name,
    }


@main_bp.before_request
def enforce_finance_page_access():
    """Finance users can only access rebate extraction pages."""
    user = get_current_user()
    if not user:
        return None

    g.current_user = user
    if not _is_finance_user(user):
        return None

    endpoint = request.endpoint
    if endpoint in _FINANCE_ALLOWED_ENDPOINTS:
        return None

    return redirect(url_for('main.rebate_extraction'))


_REBATE_ALLOC_COLUMNS = [
    ('MMC', 'MMC_ALLOC'),
    ('BURKE', 'BURKE_ALLOC'),
    ('AECOM', 'AECOM_ALLOC'),
    ('MMVO', 'MOUNT_VERNON_ALLOC'),
    ('MSSO', 'NEW_ROCHELLE_ALLOC'),
    ('NYACK', 'NYACK_ALLOC'),
    ('SLCH', 'SLCH_ALLOC'),
    ('WPH', 'WPH_ALLOC'),
]

_REBATE_ALLOC_CODE_MAP = {
    'MMC': 'MMC_ALLOC',
    'BURKE': 'BURKE_ALLOC',
    'AECOM': 'AECOM_ALLOC',
    'MMVO': 'MOUNT_VERNON_ALLOC',
    'MOUNT_VERNON': 'MOUNT_VERNON_ALLOC',
    'MOUNT VERNON': 'MOUNT_VERNON_ALLOC',
    'MSSO': 'NEW_ROCHELLE_ALLOC',
    'NEW_ROCHELLE': 'NEW_ROCHELLE_ALLOC',
    'NEW ROCHELLE': 'NEW_ROCHELLE_ALLOC',
    'NYACK': 'NYACK_ALLOC',
    'SLCH': 'SLCH_ALLOC',
    'WPH': 'WPH_ALLOC',
}

_REBATE_FILE_PATH_HEADERS = [f'FILE_PATH_{index}' for index in range(1, 11)]


def _parse_filter_date(value):
    """Parse a YYYY-MM-DD date filter value."""
    if not value:
        return None
    try:
        return datetime.strptime(value, '%Y-%m-%d').date()
    except ValueError:
        return None


def _get_template_current_user_name():
    """Return the current user's display name for templates."""
    user = getattr(g, 'current_user', None)
    if user is None:
        try:
            user = get_current_user()
        except Exception:
            user = None
    if not user:
        return None
    return user.full_name or user.username


def _normalize_role_name(user):
    """Return current user role as lowercase normalized text."""
    return (getattr(getattr(user, 'role', None), 'name', '') or '').strip().lower()


def _is_read_only_user(user):
    """Return True when the current user has a read-only role."""
    return _normalize_role_name(user) in {'read-only', 'read only', 'readonly'}


def _get_role_update_admin_email():
    """Return the mailbox users should contact for role updates."""
    return (
        current_app.config.get('REVIEW_NOTIFICATION_TO')
        or current_app.config.get('PROCUREMENT_DATA_TEAM_EMAIL')
        or current_app.config.get('FROM_EMAIL')
        or 'procurementdatateam@montefiore.org'
    )


def _get_role_update_cc_email():
    """Return the fixed CC mailbox for role update requests."""
    return (
        current_app.config.get('PROCUREMENT_DATA_TEAM_EMAIL')
        or current_app.config.get('FROM_EMAIL')
        or 'procurementdatateam@montefiore.org'
    )


def _build_role_request_mailto(user):
    """Build a prefilled mailto link to request role access changes."""
    to_email = _get_role_update_admin_email()
    cc_email = _get_role_update_cc_email()
    username = getattr(user, 'username', None) or 'unknown-user'
    full_name = getattr(user, 'full_name', None) or username
    subject = quote('Savings Tracker Role Update Request')
    body = quote(
        f"Hello Admin,\r\n\r\n"
        f"Please update my Savings Tracker role.\r\n"
        f"User: {full_name} ({username})\r\n\r\n"
        f"Requested role:\r\n"
    )
    cc = quote(cc_email)
    return f'mailto:{to_email}?cc={cc}&subject={subject}&body={body}'


def _get_allocation_numeric_value(allocation, rebate_amount=0):
    """Return an allocation as a numeric amount for export formatting."""
    if allocation.allocation_amount is not None:
        return round(float(allocation.allocation_amount), 2)
    if allocation.allocation_percentage is not None:
        return round(float(rebate_amount or 0) * float(allocation.allocation_percentage) / 100, 2)
    return None


def _format_currency_display(value):
    """Format a numeric value for display on the extraction page."""
    if value is None:
        return '—'
    return f"${float(value):,.2f}"


def _get_rebate_extraction_data(start_date=None, end_date=None, search_term=''):
    """Build rebate extraction rows, optionally filtered by check date and search text."""
    query = (
        Initiative.query.join(Initiative.rebate)
        .options(
            joinedload(Initiative.rebate),
            joinedload(Initiative.facility_allocations).joinedload(FacilityAllocation.facility),
            joinedload(Initiative.files)
        )
        .filter(
            Initiative.initiative_type == 'Rebate',
            Initiative.is_deleted == False
        )
    )

    if start_date:
        query = query.filter(Rebate.rebate_check_date >= start_date)
    if end_date:
        query = query.filter(Rebate.rebate_check_date <= end_date)

    initiatives = query.order_by(Initiative.id.desc()).all()

    rebate_rows = []
    total_rebate_amount = 0.0
    normalized_search = (search_term or '').strip().lower()

    for initiative in initiatives:
        rebate = initiative.rebate
        if rebate is None:
            continue

        rebate_amount = float(rebate.rebate_amount or 0)
        allocation_map = {column_name: None for _, column_name in _REBATE_ALLOC_COLUMNS}
        for allocation in sorted(
            initiative.facility_allocations,
            key=lambda item: ((item.facility.code if item.facility else 'ZZZ'), item.id)
        ):
            facility_code = allocation.facility.code if allocation.facility else 'Unknown'
            column_name = _REBATE_ALLOC_CODE_MAP.get(str(facility_code).upper())
            if not column_name:
                continue
            allocation_map[column_name] = _get_allocation_numeric_value(allocation, rebate_amount)

        attachments = []
        for file_record in initiative.files:
            if getattr(file_record, 'is_deleted', False):
                continue
            attachments.append({
                'id': file_record.id,
                'file_name': file_record.file_name or os.path.basename(file_record.file_path or ''),
                'file_path': file_record.file_path,
            })

        allocation_display_map = {
            column_name: _format_currency_display(value)
            for column_name, value in allocation_map.items()
        }
        row = {
            'initiative_id': initiative.id,
            'description': initiative.description or '—',
            'rebate_type': rebate.rebate_type or '—',
            'contract_number': rebate.contract_number or '—',
            'contract_category': rebate.contract_category or '—',
            'wave_category': rebate.wave_category or '—',
            'contract_source': rebate.contract_source or '—',
            'vendor_name': rebate.vendor_name or '—',
            'rebate_check_date': rebate.rebate_check_date.strftime('%Y-%m-%d') if rebate.rebate_check_date else '—',
            'rebate_payment_type': rebate.rebate_payment_type or '—',
            'check_number': rebate.check_number or '—',
            'rebate_amount': rebate_amount,
            'allocation_map': allocation_map,
            'allocation_display_map': allocation_display_map,
            'attachments': attachments,
        }

        if normalized_search:
            searchable_text = ' '.join([
                str(row['initiative_id']),
                row['description'],
                row['rebate_type'],
                row['contract_number'],
                row['contract_category'],
                row['wave_category'],
                row['contract_source'],
                row['vendor_name'],
                row['rebate_check_date'],
                row['rebate_payment_type'],
                row['check_number'],
                *allocation_display_map.values(),
            ]).lower()
            if normalized_search not in searchable_text:
                continue

        total_rebate_amount += rebate_amount
        rebate_rows.append(row)

    return rebate_rows, total_rebate_amount


def _build_rebate_excel_workbook(rebate_rows, allocation_headers, attachments_folder='Rebate_Attachments'):
    """Create the rebate extraction Excel workbook matching the reference workbook style."""
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = f"Rebate_{datetime.now().strftime('%m.%d.%y')}"
    worksheet.sheet_properties.tabColor = '92D050'

    display_headers = [
        'InitiativeID',
        'Initiative_Desc',
        'Rebates_Type',
        'Contract_Number',
        'Contract_Category',
        'Wave_Category',
        'Contract_Source',
        'Vendor_Name',
        'Rebate_Check_Date',
        'Rebate_Payment_Type',
        'Check_Number',
        'Rebate_amount',
        *allocation_headers,
        'ParentFolder',
    ]
    for file_path_header in _REBATE_FILE_PATH_HEADERS:
        suffix = file_path_header.split('_')[-1]
        display_headers.extend([f'FILE_NAME_{suffix}', file_path_header])

    worksheet.append(display_headers)
    accounting_format = r'_("$"* #,##0.00_);_("$"* \(#,##0.00\);_("$"* "-"??_);_(@_)'
    default_font = Font(name='Aptos Narrow', size=11)
    header_font = Font(name='Aptos Narrow', size=11, color='FFFFFF')
    rebate_font = Font(name='Aptos Narrow', size=11, color='9C5700')
    alloc_font = Font(name='Aptos Narrow', size=11, bold=True, color='FA7D00')
    hyperlink_font = Font(name='Aptos Narrow', size=11, color='0000FF', underline='single')
    rebate_fill = PatternFill(fill_type='solid', fgColor='FFEB9C')
    alloc_fill = PatternFill(fill_type='solid', fgColor='F2F2F2')

    column_widths = {
        'A': 12.57, 'B': 15.71, 'C': 14.86, 'D': 18.29, 'E': 18.86,
        'F': 17.29, 'G': 15.14, 'H': 20.14, 'I': 22.29, 'J': 16.29,
        'K': 16.43, 'L': 16.29, 'M': 15.14, 'N': 15.71, 'O': 15.00,
        'P': 14.57, 'Q': 15.43, 'R': 14.14, 'S': 13.71,
    }

    for col_letter, width in column_widths.items():
        worksheet.column_dimensions[col_letter].width = width

    for col_idx in range(1, len(display_headers) + 1):
        worksheet.cell(row=1, column=col_idx).font = header_font

    parent_folder_col = 12 + len(allocation_headers) + 1
    worksheet.column_dimensions[get_column_letter(parent_folder_col)].width = 19.86
    worksheet.column_dimensions[get_column_letter(parent_folder_col)].hidden = True

    for idx in range(len(_REBATE_FILE_PATH_HEADERS)):
        file_name_col = parent_folder_col + 1 + (idx * 2)
        file_path_col = file_name_col + 1
        worksheet.column_dimensions[get_column_letter(file_name_col)].width = 73
        worksheet.column_dimensions[get_column_letter(file_name_col)].hidden = True
        worksheet.column_dimensions[get_column_letter(file_path_col)].width = 73

    for row_idx, row in enumerate(rebate_rows, start=2):
        file_links = list(row.get('excel_file_links', []))[:len(_REBATE_FILE_PATH_HEADERS)]
        file_links += [None] * (len(_REBATE_FILE_PATH_HEADERS) - len(file_links))

        parent_folder = ''
        file_names = []
        for link_info in file_links:
            if link_info:
                parent_folder = parent_folder or link_info.get('parent_folder', attachments_folder)
                file_names.append(link_info.get('file_name') or '')
            else:
                file_names.append('')

        row_values = [
            row['initiative_id'],
            row['description'],
            row['rebate_type'],
            row['contract_number'],
            row['contract_category'],
            row['wave_category'],
            row['contract_source'],
            row['vendor_name'],
            row['rebate_check_date'],
            row['rebate_payment_type'],
            row['check_number'],
            row['rebate_amount'],
            *[row['allocation_map'][column_name] for column_name in allocation_headers],
            parent_folder,
        ]

        for file_name in file_names:
            row_values.extend([file_name, ''])

        worksheet.append(row_values)

        for col_idx in range(1, worksheet.max_column + 1):
            worksheet.cell(row=row_idx, column=col_idx).font = default_font

        rebate_amount_cell = worksheet.cell(row=row_idx, column=12)
        rebate_amount_cell.number_format = accounting_format
        rebate_amount_cell.fill = rebate_fill
        rebate_amount_cell.font = rebate_font

        for col_idx in range(13, 13 + len(allocation_headers)):
            alloc_cell = worksheet.cell(row=row_idx, column=col_idx)
            alloc_cell.number_format = accounting_format
            alloc_cell.fill = alloc_fill
            alloc_cell.font = alloc_font

        for idx in range(len(_REBATE_FILE_PATH_HEADERS)):
            file_name_col = parent_folder_col + 1 + (idx * 2)
            file_path_col = file_name_col + 1
            file_name = worksheet.cell(row=row_idx, column=file_name_col).value or ''
            path_cell = worksheet.cell(row=row_idx, column=file_path_col)

            if file_name:
                file_path = f'{parent_folder}\\{row["initiative_id"]}\\{file_name}'
                safe_file_path = str(file_path).replace('"', '""')
                safe_file_name = str(file_name).replace('"', '""')
                path_cell.value = f'=HYPERLINK("{safe_file_path}", "{safe_file_name}")'
            else:
                path_cell.value = ''

            path_cell.font = hyperlink_font

    table_ref = f"A1:{get_column_letter(worksheet.max_column)}{max(worksheet.max_row, 1)}"
    table = Table(displayName='RebateExtractTable', ref=table_ref)
    table.tableStyleInfo = TableStyleInfo(
        name='TableStyleLight9',
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    worksheet.add_table(table)

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


@main_bp.route('/')
def index():
    """Redirect to dashboard."""
    return redirect(url_for('main.dashboard'))


@main_bp.route('/dashboard')
@login_required
def dashboard():
    """Serve the dashboard page."""
    user = g.current_user
    return render_template(
        'dashboard.html',
        current_user=_get_template_current_user_name(),
        is_readonly_user=_is_read_only_user(user),
        role_request_mailto=_build_role_request_mailto(user),
    )


@main_bp.route('/savings-dashboard')
@login_required
def savings_dashboard():
    """Serve the savings analytics dashboard (reviewer / admin)."""
    if _is_read_only_user(g.current_user):
        flash('Read-only users can access summary information only.', 'warning')
        return redirect(url_for('main.dashboard'))
    return render_template('savings_dashboard.html', current_user=_get_template_current_user_name())


@main_bp.route('/cost-savings/form')
@login_required
def cost_savings_form():
    """Serve the cost savings form page."""
    if _is_read_only_user(g.current_user):
        flash('Read-only users cannot access detailed forms. Use the role request email button to request access.', 'warning')
        return redirect(url_for('main.dashboard'))
    return render_template('cost_savings_form.html', current_user=_get_template_current_user_name())


@main_bp.route('/rebate/form')
@login_required
def rebate_form():
    """Serve the rebate form page."""
    if _is_read_only_user(g.current_user):
        flash('Read-only users cannot access detailed forms. Use the role request email button to request access.', 'warning')
        return redirect(url_for('main.dashboard'))
    return render_template('rebate_form.html', current_user=_get_template_current_user_name())


@main_bp.route('/rebate/extraction')
@login_required
def rebate_extraction():
    """Display all rebate initiatives with their detailed facility allocations."""
    user = g.current_user
    if not (_is_admin_user(user) or _is_finance_user(user)):
        flash('Admin or Finance access is required for rebate extraction.', 'error')
        return redirect(url_for('main.dashboard'))

    start_date_raw = (request.args.get('start_date') or '').strip()
    end_date_raw = (request.args.get('end_date') or '').strip()
    search_term = (request.args.get('search') or '').strip()

    start_date = _parse_filter_date(start_date_raw)
    end_date = _parse_filter_date(end_date_raw)
    rebate_rows, total_rebate_amount = _get_rebate_extraction_data(start_date, end_date, search_term)

    return render_template(
        'rebate_extraction.html',
        current_user=_get_template_current_user_name(),
        rebate_rows=rebate_rows,
        rebate_count=len(rebate_rows),
        total_rebate_amount=total_rebate_amount,
        allocation_columns=[column_name for _, column_name in _REBATE_ALLOC_COLUMNS],
        filter_start_date=start_date_raw if start_date else '',
        filter_end_date=end_date_raw if end_date else '',
        search_term=search_term,
    )


@main_bp.route('/rebate/extraction/export')
@login_required
def rebate_extraction_export():
    """Export rebate extraction results as a ZIP containing an Excel workbook and attachments."""
    user = g.current_user
    if not (_is_admin_user(user) or _is_finance_user(user)):
        flash('Admin or Finance access is required for rebate extraction export.', 'error')
        return redirect(url_for('main.dashboard'))

    start_date = _parse_filter_date((request.args.get('start_date') or '').strip())
    end_date = _parse_filter_date((request.args.get('end_date') or '').strip())
    search_term = (request.args.get('search') or '').strip()

    rebate_rows, _ = _get_rebate_extraction_data(start_date, end_date, search_term)
    allocation_headers = [column_name for _, column_name in _REBATE_ALLOC_COLUMNS]
    attachment_folder = current_app.config.get('REBATE_ATTACHMENTS_FOLDER', 'Rebate_Attachments')
    export_date = datetime.now().strftime('%m.%d.%Y')
    workbook_name = f'Rebate Initiatives-Savings Tracker - {export_date}.xlsx'
    zip_name = f'Rebate Initiatives-Savings Tracker - {export_date}.zip'

    zip_buffer = io.BytesIO()
    missing_files = []

    with zipfile.ZipFile(zip_buffer, 'w', compression=zipfile.ZIP_DEFLATED) as zip_file:
        attachment_count = 0
        used_archive_names = set()

        for row in rebate_rows:
            excel_file_links = []
            for attachment in row['attachments']:
                file_path = attachment.get('file_path')
                original_name = os.path.basename(attachment.get('file_name') or '') or f"file_{attachment.get('id', 'unknown')}"

                if file_path and os.path.isfile(file_path):
                    attachment_count += 1
                    file_root, file_ext = os.path.splitext(original_name)
                    archive_file_name = original_name
                    if archive_file_name.lower() in used_archive_names:
                        archive_file_name = f"{file_root}_{attachment.get('id', 'file')}{file_ext}"
                    used_archive_names.add(archive_file_name.lower())

                    arcname = f"{attachment_folder}/{row['initiative_id']}/{archive_file_name}"
                    zip_file.write(file_path, arcname=arcname)
                    if len(excel_file_links) < len(_REBATE_FILE_PATH_HEADERS):
                        excel_file_links.append({
                            'parent_folder': attachment_folder,
                            'file_name': archive_file_name,
                        })
                else:
                    missing_files.append(
                        f"Initiative {row['initiative_id']}: {attachment.get('file_name', 'Unknown file')}"
                    )

            row['excel_file_links'] = excel_file_links

        workbook_buffer = _build_rebate_excel_workbook(rebate_rows, allocation_headers, attachment_folder)
        zip_file.writestr(workbook_name, workbook_buffer.getvalue())

        if attachment_count == 0:
            zip_file.writestr(
                f'{attachment_folder}/README.txt',
                'No attachment files were found for the selected rebate initiatives.\n'
            )

        if missing_files:
            zip_file.writestr(f'{attachment_folder}/MISSING_FILES.txt', '\n'.join(missing_files))

    zip_buffer.seek(0)
    return send_file(
        zip_buffer,
        mimetype='application/zip',
        as_attachment=True,
        download_name=zip_name
    )


@main_bp.route('/cost-avoidance/form')
@login_required
def cost_avoidance_form():
    """Serve the cost avoidance form page."""
    if _is_read_only_user(g.current_user):
        flash('Read-only users cannot access detailed forms. Use the role request email button to request access.', 'warning')
        return redirect(url_for('main.dashboard'))
    return render_template('cost_avoidance_form.html', current_user=_get_template_current_user_name())


def _is_admin_user(user):
    """Return True when the current user is an administrator."""
    if not user:
        return False
    return any([
        user.has_permission('manage_users'),
        getattr(getattr(user, 'role', None), 'name', '') == 'Admin',
    ])


def _is_finance_user(user):
    """Return True when the current user has the Finance role."""
    if not user:
        return False
    return getattr(getattr(user, 'role', None), 'name', '') == 'Finance'


def _can_access_email_testing(user):
    """Return True when the current user should see the admin email test tools."""
    return _is_admin_user(user)


def _normalize_email_list(values):
    """Return a de-duplicated list of trimmed email addresses."""
    normalized = []
    seen = set()
    for value in values or []:
        address = (value or '').strip()
        if not address:
            continue
        key = address.lower()
        if key in seen:
            continue
        seen.add(key)
        normalized.append(address)
    return normalized


def _get_review_notification_options():
    """Return the allowed recipients for review notification settings."""
    eligible_users = (
        User.query
        .join(UserRole)
        .filter(
            User.is_active == True,
            UserRole.name.in_(['Reviewer', 'Admin']),
        )
        .order_by(UserRole.name.asc(), User.full_name.asc())
        .all()
    )

    options = [{
        'value': _REVIEW_NOTIFICATION_DEFAULT_EMAIL,
        'label': f'{_REVIEW_NOTIFICATION_DEFAULT_EMAIL} (default)',
    }]

    seen = {_REVIEW_NOTIFICATION_DEFAULT_EMAIL.lower()}
    for user in eligible_users:
        email = (user.email or '').strip()
        if not email:
            continue
        email_key = email.lower()
        if email_key in seen:
            continue
        seen.add(email_key)
        role_name = getattr(getattr(user, 'role', None), 'name', '') or 'User'
        options.append({
            'value': email,
            'label': f'{user.full_name} <{email}> ({role_name})',
        })

    return options


@main_bp.route('/admin/email-notifications')
@login_required
def email_notifications_admin():
    """Admin page for testing Microsoft Graph email notifications."""
    user = g.current_user
    if not _can_access_email_testing(user):
        flash('Admin access is required to test email notifications.', 'error')
        return redirect(url_for('main.dashboard'))

    initiatives = (
        Initiative.query
        .filter_by(is_deleted=False)
        .order_by(Initiative.id.desc())
        .limit(50)
        .all()
    )

    current_settings = current_app.config.get('STATIC_APP_SETTINGS') or load_static_settings(current_app.config)
    review_notification_options = _get_review_notification_options()
    weekly_reminder_options = review_notification_options
    stored_review_recipients = _normalize_email_list(
        (current_settings.get('email', {}) or {}).get('review_notification_to', '').replace(';', ',').split(',')
    )
    stored_review_recipient_keys = {email.lower() for email in stored_review_recipients}
    stored_weekly_recipients = _normalize_email_list(
        (current_settings.get('email', {}) or {}).get('weekly_reminder_to', '').replace(';', ',').split(',')
    )
    stored_weekly_recipient_keys = {email.lower() for email in stored_weekly_recipients}

    return render_template(
        'email_notifications.html',
        current_user=user.full_name,
        initiatives=initiatives,
        graph_ready=is_graph_mail_configured(),
        sender_mailbox=current_app.config.get('MS_GRAPH_SENDER_USER_ID') or current_app.config.get('FROM_EMAIL'),
        review_mailbox=current_app.config.get('REVIEW_NOTIFICATION_TO') or current_app.config.get('PROCUREMENT_DATA_TEAM_EMAIL'),
        procurement_mailbox=current_app.config.get('PROCUREMENT_DATA_TEAM_EMAIL'),
        selected_id=request.args.get('selected_id', type=int),
        settings=current_settings,
        settings_file_path=current_app.config.get('STATIC_CONFIG_FILE'),
        review_notification_options=review_notification_options,
        review_notification_selected=[
            option['value']
            for option in review_notification_options
            if option['value'].lower() in stored_review_recipient_keys
        ] or [_REVIEW_NOTIFICATION_DEFAULT_EMAIL],
        weekly_reminder_options=weekly_reminder_options,
        weekly_reminder_selected=[
            option['value']
            for option in weekly_reminder_options
            if option['value'].lower() in stored_weekly_recipient_keys
        ] or [_REVIEW_NOTIFICATION_DEFAULT_EMAIL],
    )


@main_bp.route('/admin/email-notifications/send', methods=['POST'])
@login_required
def send_test_email_notification():
    """Send one of the configured email notifications for admin testing."""
    user = g.current_user
    if not _can_access_email_testing(user):
        flash('Admin access is required to test email notifications.', 'error')
        return redirect(url_for('main.dashboard'))

    email_type = (request.form.get('email_type') or '').strip().lower()
    initiative_id = request.form.get('initiative_id', type=int)
    initiative = None

    if initiative_id:
        initiative = Initiative.query.filter_by(id=initiative_id, is_deleted=False).first()

    if email_type != 'weekly' and not initiative:
        flash('Please select a valid initiative for the test email.', 'error')
        return redirect(url_for('main.email_notifications_admin', selected_id=initiative_id or ''))

    reviewers = User.query.join(UserRole).filter(
        UserRole.can_review == True,
        User.is_active == True,
    ).all()

    success = False
    label = 'email notification'
    email_result = {'success': False, 'message': 'No email result was returned.', 'status_code': None}

    if email_type == 'created':
        email_result = send_initiative_creator_notification(initiative, initiative.creator or user, return_details=True)
        label = f'new initiative created email for #{initiative.id}'
    elif email_type == 'review':
        email_result = send_initiative_review_notification(initiative, initiative.creator or user, reviewers, return_details=True)
        label = f'review notification email for #{initiative.id}'
    elif email_type == 'approved':
        if initiative.review_date is None:
            initiative.review_date = initiative.updated_at or initiative.created_at or now_eastern()
        email_result = send_initiative_approved_notification(initiative, initiative.reviewer or user, return_details=True)
        label = f'approval email for #{initiative.id}'
    elif email_type == 'weekly':
        email_result = send_weekly_review_reminder(return_details=True)
        label = 'weekly reminder email'
    else:
        flash('Unknown email notification type selected.', 'error')
        return redirect(url_for('main.email_notifications_admin', selected_id=initiative_id or ''))

    success = bool(email_result.get('success'))
    if success:
        flash(f"Successfully sent the {label}. {email_result.get('message', '')}".strip(), 'success')
    else:
        error_message = email_result.get('message', 'Unknown Microsoft Graph error.')
        status_code = email_result.get('status_code')
        status_suffix = f' (status {status_code})' if status_code else ''
        flash(f'Unable to send the {label}{status_suffix}: {error_message}', 'warning')

    return redirect(url_for('main.email_notifications_admin', selected_id=initiative_id or ''))


@main_bp.route('/admin/static-config/save', methods=['POST'])
@login_required
def save_static_config():
    """Save admin-managed notification and file path settings to the static config file."""
    user = g.current_user
    if not _can_access_email_testing(user):
        flash('Admin access is required to update settings.', 'error')
        return redirect(url_for('main.dashboard'))

    fixed_mailbox = _REVIEW_NOTIFICATION_DEFAULT_EMAIL
    allowed_review_addresses = {
        option['value'].lower(): option['value']
        for option in _get_review_notification_options()
    }
    selected_review_addresses = []
    for address in request.form.getlist('review_notification_to'):
        normalized_address = (address or '').strip()
        if not normalized_address:
            continue
        allowed_value = allowed_review_addresses.get(normalized_address.lower())
        if allowed_value:
            selected_review_addresses.append(allowed_value)
    if not selected_review_addresses:
        selected_review_addresses = [fixed_mailbox]

    selected_weekly_addresses = []
    for address in request.form.getlist('weekly_reminder_to'):
        normalized_address = (address or '').strip()
        if not normalized_address:
            continue
        allowed_value = allowed_review_addresses.get(normalized_address.lower())
        if allowed_value:
            selected_weekly_addresses.append(allowed_value)
    if not selected_weekly_addresses:
        selected_weekly_addresses = [fixed_mailbox]

    settings = load_static_settings(current_app.config)
    settings['email'] = {
        'from_email': fixed_mailbox,
        'graph_sender_user_id': fixed_mailbox,
        'creator_notification_to': (request.form.get('creator_notification_to') or '').strip(),
        'review_notification_to': ', '.join(_normalize_email_list(selected_review_addresses)),
        'approval_notification_to': (request.form.get('approval_notification_to') or '').strip(),
        'weekly_reminder_to': ', '.join(_normalize_email_list(selected_weekly_addresses)),
        'cc_addresses': fixed_mailbox,
    }
    settings['files'] = {
        'file_storage_path': (request.form.get('file_storage_path') or '').strip(),
        'uploads_fallback_path': (request.form.get('uploads_fallback_path') or '').strip() or 'uploads',
        'rebate_attachments_folder': (request.form.get('rebate_attachments_folder') or '').strip() or 'Rebate_Attachments',
        'logs_path': (request.form.get('logs_path') or '').strip() or 'logs',
    }

    save_static_settings(settings)
    apply_static_settings(current_app)
    flash(
        f'Static configuration updated successfully. Review notification and weekly reminder recipients are limited to {fixed_mailbox} and active Reviewer/Admin users. Sender and CC remain fixed to {fixed_mailbox}.',
        'success'
    )
    return redirect(url_for('main.email_notifications_admin'))


@main_bp.route('/admin/users')
@login_required
def user_management():
    """Admin page for managing user accounts and role assignments."""
    user = g.current_user
    if not _is_admin_user(user):
        flash('Admin access is required to manage users.', 'error')
        return redirect(url_for('main.dashboard'))

    users = User.query.order_by(User.full_name).all()
    roles = UserRole.query.order_by(UserRole.name).all()

    return render_template(
        'user_management.html',
        current_user=user.full_name,
        current_user_id=user.id,
        users=users,
        roles=roles,
    )


@main_bp.route('/admin/users/<int:user_id>/role', methods=['POST'])
@login_required
def user_assign_role(user_id):
    """Assign a role to a user."""
    admin = g.current_user
    if not _is_admin_user(admin):
        flash('Admin access required.', 'error')
        return redirect(url_for('main.dashboard'))

    target = User.query.get_or_404(user_id)
    role_id = request.form.get('role_id', type=int)
    role = UserRole.query.get(role_id) if role_id else None

    if not role:
        flash('Invalid role selected.', 'error')
        return redirect(url_for('main.user_management'))

    from app import db
    old_role = target.role.name if target.role else 'None'
    target.role_id = role.id
    db.session.commit()
    flash(f'Role updated for {target.full_name or target.username}: {old_role} → {role.name}.', 'success')
    return redirect(url_for('main.user_management'))


@main_bp.route('/admin/users/<int:user_id>/toggle-status', methods=['POST'])
@login_required
def user_toggle_status(user_id):
    """Activate or deactivate a user account."""
    admin = g.current_user
    if not _is_admin_user(admin):
        flash('Admin access required.', 'error')
        return redirect(url_for('main.dashboard'))

    if user_id == admin.id:
        flash('You cannot change the status of your own account.', 'error')
        return redirect(url_for('main.user_management'))

    target = User.query.get_or_404(user_id)
    from app import db
    target.is_active = not target.is_active
    db.session.commit()
    status_label = 'activated' if target.is_active else 'deactivated'
    flash(f'User {target.full_name or target.username} has been {status_label}.', 'success')
    return redirect(url_for('main.user_management'))


@main_bp.route('/admin/users/<int:user_id>/delete', methods=['POST'])
@login_required
def user_delete(user_id):
    """Permanently delete a user account."""
    admin = g.current_user
    if not _is_admin_user(admin):
        flash('Admin access required.', 'error')
        return redirect(url_for('main.dashboard'))

    if user_id == admin.id:
        flash('You cannot delete your own account.', 'error')
        return redirect(url_for('main.user_management'))

    target = User.query.get_or_404(user_id)
    display_name = target.full_name or target.username
    from app import db
    db.session.delete(target)
    db.session.commit()
    flash(f'User {display_name} has been permanently deleted.', 'success')
    return redirect(url_for('main.user_management'))


@main_bp.route('/admin/distribution')
@login_required
def admin_distribution():
    """Admin page for running the daily distribution stored procedure."""
    user = g.current_user
    if not _is_admin_user(user):
        flash('Admin access is required to run the PowerBI dashboard backend force refresh.', 'error')
        return redirect(url_for('main.dashboard'))
    return render_template('admin_distribution.html', current_user=user.full_name)


@main_bp.route('/logout')
def logout():
    """Handle logout."""
    # For IIS Windows Authentication, we can't really logout
    # Just redirect to a goodbye page or back to index
    return '''
    <!DOCTYPE html>
    <html>
    <head>
        <title>Logged Out - Savings Tracker</title>
        <meta charset="utf-8">
        <meta name="viewport" content="width=device-width, initial-scale=1">
        <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/css/bootstrap.min.css" rel="stylesheet">
    </head>
    <body class="bg-light">
        <div class="container mt-5">
            <div class="row justify-content-center">
                <div class="col-md-6 text-center">
                    <h1>Logged Out</h1>
                    <p class="lead">You have been logged out. Close your browser to complete the logout process.</p>
                    <a href="/" class="btn btn-primary">Return to Login</a>
                </div>
            </div>
        </div>
    </body>
    </html>
    '''


@main_bp.route('/static/<path:path>')
def send_static(path):
    """Serve static files."""
    return send_from_directory('static', path)
